#!/usr/bin/env python
"""Static and release-integrity checks for the public repository."""
from __future__ import annotations

import argparse
import ast
import csv
import hashlib
import math
import re
import subprocess
import sys
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
CODE_SUFFIXES = {".py", ".R", ".r", ".ps1", ".md"}
FORBIDDEN = (
    re.compile(r"[A-Za-z]:[\\/](?:Users|onfh-pilot|ONFH2026dry)[\\/]", re.I),
    re.compile(r"/home/[^/]+/"),
)

REQUIRED = [
    "README.md",
    ".python-version",
    "analysis/sample_metadata_v4.csv",
    "analysis/v4_common.py",
    "analysis/prepare_matrices.R",
    "analysis/participant_fgsea_stability.py",
    "analysis/participant_fgsea_stability.R",
    "analysis/serum_paired_comparison.R",
    "analysis/spatial_contextualization.py",
    "virtual_knockout/run_official_vko.R",
    "virtual_knockout/run_matched_control_vko.R",
    "virtual_knockout/postprocess_mt_exclusion.R",
    "virtual_knockout/vko_selected_features_v5.csv",
    "virtual_knockout/vko_input_checksums.tsv",
    "plotting/assemble_manuscript_figures.py",
    "plotting/make_reviewed_figures.py",
    "plotting/make_figure4.py",
    "plotting/make_virtual_knockout_figure.R",
    "plotting/make_genes_revision_figures.R",
    "plotting/make_genes_virtual_knockout_figure.R",
    "plotting/compress_spatial_figure_for_portal.R",
    "environment/r-package-versions.tsv",
    "environment/check_r_packages.R",
    "results/README.md",
    "results/figure_inputs/figure2_umap.csv.gz",
    "results/figure_inputs/figure2_marker_dotplot.csv",
    "results/figure_inputs/ec_subtype_composition_v4.csv",
    "results/figure_inputs/module_scores_by_library_v4.csv",
    "results/figure_inputs/module_scores_liao_stats_v4.csv",
    "results/figure_inputs/de_ec_SONFH_vs_HOA_descriptive_v4.csv",
    "results/figure_inputs/fig3_key_gene_effects_v4.csv",
    "results/figure_inputs/gsea_ONFH3A_vs_HOA_H.csv",
    "results/figure_inputs/sample_level_tf_ulm_v4.csv",
    "results/figure_inputs/tf_stats_ulm_v4.csv",
    "results/figure_inputs/comm_scores_v4_long.csv.gz",
    "results/figure_inputs/comm_top_ONFH_4_vs_ONFH_3A_v4.csv",
    "results/figure_inputs/diag_summary_v4.json",
    "results/figure_inputs/diag_permutation_v4.csv",
    "results/figure_inputs/diag_feature_stability_v4.csv",
    "results/figure_inputs/diag_nested_cv_performance_v7.csv",
    "results/figure_inputs/diag_ma_comparator_repeat_performance_v7.csv",
    "results/figure_inputs/diag_oof_predictions_aggregated_v8.csv",
    "results/figure_inputs/diag_paired_model_comparison_v8.csv",
    "results/figure_inputs/diag_delong_model_comparison_v8.csv",
    "results/official_r_vko_figure_data/sqstm1_outgoing_edges_hoa2.csv",
    "results/official_r_vko_figure_data/sqstm1_outgoing_edges_hoa3.csv",
    "results/official_r_vko_figure_data/wt_ko_manifold_hoa2.csv",
    "results/official_r_vko_figure_data/wt_ko_manifold_hoa3.csv",
    "results/official_r_vko_manuscript/vko_sqstm1_hoa2_official_r.csv",
    "results/official_r_vko_manuscript/vko_sqstm1_hoa3_official_r.csv",
    "results/official_r_vko_manuscript/vko_sqstm1_pathway_enrichment_official_r.csv",
    "results/official_r_vko_cross_profile_gene_summary.csv",
    "results/official_r_vko_manuscript_no_mt_encoded/vko_provenance_official_r.json",
    "results/official_r_vko_official_default_no_mt_encoded/vko_provenance_official_r.json",
    "results/official_r_vko_no_mt_gene_comparison.csv",
    "results/official_r_vko_no_mt_summary.csv",
    "results/official_r_vko_no_mt_cross_donor_audit.csv",
    "results/official_r_vko_no_mt_hoa3_fdr_audit.csv",
    "results/official_r_vko_no_mt_pathway_by_donor.csv",
    "results/official_r_vko_no_mt_pathway_summary.csv",
    "results/vko_mt_encoded_exclusion_manifest.csv",
    "results/official_r_vko_matched_controls/matched_control_selection.csv",
    "results/official_r_vko_matched_controls/matched_control_vko_summary.csv",
    "results/official_r_vko_matched_controls/sqstm1_matched_control_calibration.csv",
    "results/official_r_vko_matched_controls/frozen_wt_reuse_validation.csv",
    "results/official_r_vko_matched_controls/sqstm1_selection_rationale.csv",
    "results/participant_fgsea_stability/fgsea_leading_edge_top20.csv",
    "results/participant_fgsea_stability/fgsea_lopo_summary.csv",
    "results/participant_fgsea_stability/fgsea_selected_pathway_overlap.csv",
    "results/spatial_contextualization/GSE284089_spatial_spot_scores.csv.gz",
    "results/spatial_contextualization/spatial_module_gene_coverage.csv",
    "results/spatial_contextualization/spatial_expression_matched_controls.csv",
    "results/spatial_contextualization/spatial_within_section_correlations.csv",
    "results/spatial_contextualization/spatial_run_summary.csv",
    "results/spatial_contextualization/spatial_provenance.json",
    "workflow/run_core_analysis.ps1",
    "workflow/run_virtual_knockout.ps1",
    "workflow/run_spatial_contextualization.ps1",
    "workflow/run_genes_revision_robustness.ps1",
    "workflow/run_figures.ps1",
    "results/Supplementary_Tables_S1-S11.xlsx",
]

DATA_REQUIRED = [
    *[f"data/liao2022/{sample}.rds" for sample in (
        "onfh1", "onfh2", "onfh3", "onfh4", "onfh5", "onfh6",
        "hoa1", "hoa2", "hoa3", "fnf1", "fnf2"
    )],
    "data/GSE123568_series_matrix.txt.gz",
    "data/GSE123568_family.soft.gz",
    "data/cellchatdb_interactions.csv",
    "data/cellchatdb_complex_named.csv",
    "data/dorothea_ABC.tsv",
    "data/gse284089/GSE284089_RAW.tar",
    "data/gse284089/GSM8677818_matrix.mtx.gz",
    "data/gse284089/GSM8677818_barcodes.tsv.gz",
    "data/gse284089/GSM8677818_features.tsv.gz",
    "data/gse284089/GSM8677818_tissue_positions.csv.gz",
    "data/gse284089/GSM8677818_scalefactors_json.json.gz",
    "data/gse284089/GSM8677818_tissue_hires_image.png.gz",
]


def fail(errors: list[str], message: str) -> None:
    errors.append(message)
    print(f"FAIL  {message}")


def pass_(message: str) -> None:
    print(f"PASS  {message}")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def versioned_files() -> list[Path]:
    """Return Git-tracked files, with a source-archive fallback."""
    completed = subprocess.run(
        ["git", "ls-files", "-z"],
        cwd=ROOT,
        capture_output=True,
        check=False,
    )
    if completed.returncode == 0 and completed.stdout:
        return [
            ROOT / raw.decode("utf-8")
            for raw in completed.stdout.split(b"\0")
            if raw
        ]
    return [path for path in ROOT.rglob("*") if path.is_file() and ".git" not in path.parts]


def same_value(actual: object, expected: object) -> bool:
    if isinstance(expected, (int, float)):
        try:
            return math.isclose(float(actual), float(expected), rel_tol=1e-11, abs_tol=1e-14)
        except (TypeError, ValueError):
            return False
    return str(actual) == str(expected)


def audit_s10g_workbook() -> list[str]:
    """Reconcile the detailed S10g and main-text S10i audits with source CSVs."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["openpyxl is unavailable; install environment/requirements-python.lock.txt"]

    workbook_path = ROOT / "results/Supplementary_Tables_S1-S11.xlsx"
    cross_path = ROOT / "results/official_r_vko_no_mt_cross_donor_audit.csv"
    gene_path = ROOT / "results/official_r_vko_no_mt_hoa3_fdr_audit.csv"
    with cross_path.open(encoding="utf-8", newline="") as handle:
        cross_rows = list(csv.DictReader(handle))
    with gene_path.open(encoding="utf-8", newline="") as handle:
        gene_rows = list(csv.DictReader(handle))

    if len(cross_rows) != 2 or [row["gene"] for row in gene_rows] != ["VWF", "NFKBIA", "C7"]:
        return ["source audit CSV row counts or HOA3 gene order changed"]

    # Artifact-authored OOXML may omit worksheet dimension metadata, which makes
    # ``max_row`` unavailable in openpyxl's read-only mode. Load normally so the
    # release audit derives worksheet dimensions from the actual cells.
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        if "Table S10" not in workbook.sheetnames:
            return ["Table S10 worksheet is missing"]
        sheet = workbook["Table S10"]
        issues: list[str] = []
        if sheet["K717"].value != "S10g-5A. Cross-donor common-nuclear rank and FDR audit":
            issues.append("Table S10!K717 title")
        if sheet["K723"].value != "S10g-5B. Primary-refit HOA3 downstream FDR-positive genes":
            issues.append("Table S10!K723 title")
        if sheet["A965"].value != "S10i. Main-text cross-donor common-nuclear audit after mtDNA-feature exclusion":
            issues.append("Table S10!A965 title")

        cross_fields = (
            "profile", "model_variant", "donor_pair", "n_common_nuclear_genes",
            "spearman_rho", "spearman_p_value", "top20_overlap_count",
            "common_bh_fdr_hoa2_count", "common_bh_fdr_hoa3_count",
            "common_bh_fdr_replicated_count",
        )
        cross_numeric = set(cross_fields[3:])
        for row_number, source_row in zip((719, 720), cross_rows):
            expected = [
                float(source_row[field]) if field in cross_numeric else source_row[field]
                for field in cross_fields
            ]
            actual = [sheet.cell(row_number, column).value for column in range(11, 21)]
            for column, (observed, wanted) in enumerate(zip(actual, expected), 11):
                if not same_value(observed, wanted):
                    issues.append(f"Table S10!{sheet.cell(row_number, column).coordinate}")

        s10i_fields = (
            "profile", "model_variant", "donor_pair", "n_common_nuclear_genes",
            "rank_metric", "spearman_rho", "spearman_p_value", "top20_overlap_count",
            "top20_jaccard", "common_bh_fdr_hoa2_count", "common_bh_fdr_hoa3_count",
            "common_bh_fdr_replicated_count", "common_bh_fdr_hoa2_genes",
            "common_bh_fdr_hoa3_genes", "common_bh_fdr_replicated_genes",
        )
        s10i_numeric = {
            "n_common_nuclear_genes", "spearman_rho", "spearman_p_value",
            "top20_overlap_count", "top20_jaccard", "common_bh_fdr_hoa2_count",
            "common_bh_fdr_hoa3_count", "common_bh_fdr_replicated_count",
        }
        for row_number, source_row in zip((967, 968), cross_rows):
            expected = [
                float(source_row[field]) if field in s10i_numeric else (source_row[field] or None)
                for field in s10i_fields
            ]
            actual = [sheet.cell(row_number, column).value for column in range(1, 16)]
            for column, (observed, wanted) in enumerate(zip(actual, expected), 1):
                if not same_value(observed, wanted):
                    issues.append(f"Table S10!{sheet.cell(row_number, column).coordinate}")

        gene_fields = (
            "gene", "official_rank", "common_rank", "common_rank_percentile", "distance",
            "z_score", "fold_change", "raw_p_value", "official_bh_fdr_295_family",
            "common_nuclear_bh_fdr_294_family",
        )
        for row_number, source_row in zip((725, 726, 727), gene_rows):
            expected = [
                source_row[field] if field == "gene" else float(source_row[field])
                for field in gene_fields
            ]
            actual = [sheet.cell(row_number, column).value for column in range(11, 21)]
            for column, (observed, wanted) in enumerate(zip(actual, expected), 11):
                if not same_value(observed, wanted):
                    issues.append(f"Table S10!{sheet.cell(row_number, column).coordinate}")
        return issues
    finally:
        workbook.close()


def audit_s11_workbook() -> list[str]:
    """Check spatial summary and row counts against the versioned source tables."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["openpyxl is unavailable; install environment/requirements-python.lock.txt"]

    workbook_path = ROOT / "results/Supplementary_Tables_S1-S11.xlsx"
    with (ROOT / "results/spatial_contextualization/spatial_run_summary.csv").open(
        encoding="utf-8", newline=""
    ) as handle:
        summary = next(csv.DictReader(handle))
    with (ROOT / "results/spatial_contextualization/spatial_module_gene_coverage.csv").open(
        encoding="utf-8", newline=""
    ) as handle:
        coverage_rows = list(csv.DictReader(handle))
    with (ROOT / "results/spatial_contextualization/spatial_expression_matched_controls.csv").open(
        encoding="utf-8", newline=""
    ) as handle:
        control_rows = list(csv.DictReader(handle))

    # Artifact-authored OOXML may omit worksheet dimension metadata, which makes
    # ``max_row`` unavailable in openpyxl's read-only mode. Load normally so the
    # release audit derives worksheet dimensions from the actual cells.
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        required_sheets = {"Table S11a", "Table S11b", "Table S11c", "Table S11d"}
        if not required_sheets.issubset(workbook.sheetnames):
            return ["one or more Table S11a-S11d worksheets are missing"]
        issues: list[str] = []
        sheet = workbook["Table S11a"]
        observed = {
            sheet.cell(row, 1).value: sheet.cell(row, 2).value
            for row in range(6, 20)
        }
        for field in (
            "accession", "sample", "spots_before_qc", "spots_after_qc",
            "ec_enriched_spots", "sqstm1_detected_spots",
        ):
            if not same_value(observed.get(field), summary[field]):
                issues.append(f"Table S11a summary field {field}")
        if workbook["Table S11b"].max_row != len(coverage_rows) + 5:
            issues.append("Table S11b gene-coverage row count")
        if workbook["Table S11c"].max_row != len(control_rows) + 5:
            issues.append("Table S11c matched-control row count")
        if workbook["Table S11d"].max_row != int(summary["spots_after_qc"]) + 5:
            issues.append("Table S11d spot-score row count")
        return issues
    finally:
        workbook.close()


def audit_checksum_manifest(manifest_path: Path) -> list[str]:
    """Check locally present primary inputs against a versioned SHA256 manifest."""
    with manifest_path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    issues: list[str] = []
    for row in rows:
        relative = row["relative_path"]
        path = ROOT / relative
        if not path.exists():
            issues.append(f"missing {relative}")
            continue
        if path.stat().st_size != int(row["bytes"]):
            issues.append(f"size mismatch {relative}")
        if sha256(path).lower() != row["sha256"].lower():
            issues.append(f"SHA256 mismatch {relative}")
    return issues


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check-data", action="store_true", help="also require downloaded primary data")
    parser.add_argument(
        "--check-vko-data",
        action="store_true",
        help="verify the exact HOA2/HOA3 RDS inputs used by the virtual-knockout release",
    )
    parser.add_argument(
        "--rscript",
        type=Path,
        help="optional Rscript executable used to parse every versioned R file",
    )
    parser.add_argument(
        "--submission-dir",
        type=Path,
        help="optional submission figure directory whose PDFs must byte-match figures/final",
    )
    args = parser.parse_args()
    errors: list[str] = []
    tracked = versioned_files()

    missing = [name for name in REQUIRED if not (ROOT / name).exists()]
    if missing:
        fail(errors, "missing release files: " + ", ".join(missing))
    else:
        pass_(f"{len(REQUIRED)} required release files")

    code_roots = {"analysis", "virtual_knockout", "plotting", "qa"}
    python_files = sorted(
        path for path in tracked
        if path.suffix == ".py" and path.relative_to(ROOT).parts[0] in code_roots
    )
    syntax_errors = []
    for path in python_files:
        try:
            ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        except (SyntaxError, UnicodeDecodeError) as exc:
            syntax_errors.append(f"{path.relative_to(ROOT)}: {exc}")
    if syntax_errors:
        fail(errors, "Python syntax: " + " | ".join(syntax_errors))
    else:
        pass_(f"Python syntax ({len(python_files)} files)")

    if args.rscript:
        r_roots = {"analysis", "virtual_knockout", "plotting", "environment"}
        r_files = sorted(
            path for path in tracked
            if path.suffix == ".R" and path.relative_to(ROOT).parts[0] in r_roots
        )
        r_syntax_errors = []
        for path in r_files:
            r_path = str(path).replace("\\", "/")
            completed = subprocess.run(
                [str(args.rscript), "-e", f"invisible(parse(file={r_path!r}))"],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )
            if completed.returncode != 0:
                detail = (completed.stderr or completed.stdout).strip().replace("\n", " | ")
                r_syntax_errors.append(f"{path.relative_to(ROOT)}: {detail}")
        if r_syntax_errors:
            fail(errors, "R syntax: " + " | ".join(r_syntax_errors))
        else:
            pass_(f"R syntax ({len(r_files)} files)")

    local_paths = []
    for path in tracked:
        if not path.is_file() or path.suffix not in CODE_SUFFIXES:
            continue
        if path.resolve() == Path(__file__).resolve():
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        for line_number, line in enumerate(text.splitlines(), 1):
            if any(pattern.search(line) for pattern in FORBIDDEN):
                local_paths.append(f"{path.relative_to(ROOT)}:{line_number}")
    if local_paths:
        fail(errors, "machine-specific absolute paths: " + ", ".join(local_paths))
    else:
        pass_("no machine-specific absolute paths")

    with (ROOT / "analysis/sample_metadata_v4.csv").open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    participants = {row["participant_id"] for row in rows if row["participant_id"]}
    sonfh = [row for row in rows if row["dataset"] == "sonfh_cystic"]
    metadata_ok = (
        len(rows) == 19
        and len(participants) == 15
        and len(sonfh) == 4
        and all(row["independent_for_inference"].strip().lower() in {"false", "0"} for row in sonfh)
    )
    if metadata_ok:
        pass_("sampling-unit guardrail: 19 libraries, 15 mapped participants, 4 descriptive SONFH libraries")
    else:
        fail(errors, "sampling-unit metadata no longer matches the frozen inferential design")

    with (ROOT / "results/figure_inputs/gsea_ONFH3A_vs_HOA_H.csv").open(
        encoding="utf-8", newline=""
    ) as handle:
        gsea = {row["pathway"]: float(row["padj"]) for row in csv.DictReader(handle)}
    expected_gsea = {
        "HALLMARK_ALLOGRAFT_REJECTION": 1.12899319297421e-17,
        "HALLMARK_INTERFERON_GAMMA_RESPONSE": 2.11625060949194e-14,
        "HALLMARK_INTERFERON_ALPHA_RESPONSE": 8.32810112872156e-10,
        "HALLMARK_INFLAMMATORY_RESPONSE": 1.15308484547667e-9,
    }
    if all(
        name in gsea and math.isclose(gsea[name], value, rel_tol=1e-12)
        for name, value in expected_gsea.items()
    ):
        pass_("Figure 4A GSEA FDR values match the frozen manuscript results")
    else:
        fail(errors, "Figure 4A GSEA table no longer matches the frozen manuscript results")

    s10g_issues = audit_s10g_workbook()
    if s10g_issues:
        fail(errors, "Table S10g workbook/CSV mismatch: " + ", ".join(s10g_issues))
    else:
        pass_("Table S10g/S10i audit cells match the versioned cross-donor and HOA3 CSVs")

    s11_issues = audit_s11_workbook()
    if s11_issues:
        fail(errors, "Table S11 workbook/source mismatch: " + ", ".join(s11_issues))
    else:
        pass_("Table S11 spatial summary and row counts match the versioned source tables")

    final_expected = [f"figures/final/Figure{i}.{ext}" for i in range(1, 8) for ext in ("pdf", "png")]
    final_expected += [f"figures/final/SupplementaryFigureS1.{ext}" for ext in ("pdf", "png")]
    final_missing = [name for name in final_expected if not (ROOT / name).exists()]
    if final_missing:
        fail(errors, "missing final figures: " + ", ".join(final_missing))
    else:
        pass_("all submission figures are present as PDF and PNG")

    if args.submission_dir:
        names = [f"Figure{i}.pdf" for i in range(1, 8)]
        submission_names = {name: name for name in names}
        supplement_candidates = ("SupplementaryFigureS1.pdf", "Supplementary_Figure_S1.pdf")
        supplement_name = next(
            (name for name in supplement_candidates if (args.submission_dir / name).exists()),
            supplement_candidates[0],
        )
        submission_names["SupplementaryFigureS1.pdf"] = supplement_name
        missing_submission = [
            submitted
            for submitted in submission_names.values()
            if not (args.submission_dir / submitted).exists()
        ]
        if missing_submission:
            fail(errors, "missing PDFs in --submission-dir: " + ", ".join(missing_submission))
        else:
            mismatched = [
                canonical
                for canonical, submitted in submission_names.items()
                if sha256(ROOT / "figures" / "final" / canonical)
                != sha256(args.submission_dir / submitted)
            ]
            if mismatched:
                fail(errors, "repository/submission figure hash mismatch: " + ", ".join(mismatched))
            else:
                pass_("repository and submission figure PDFs are byte-identical")

        submission_workbook = args.submission_dir / "Supplementary_Tables_S1-S11.xlsx"
        if not submission_workbook.exists():
            fail(errors, "missing Supplementary_Tables_S1-S11.xlsx in --submission-dir")
        elif sha256(ROOT / "results/Supplementary_Tables_S1-S11.xlsx") != sha256(submission_workbook):
            fail(errors, "repository/submission supplementary-workbook hash mismatch")
        else:
            pass_("repository and submission supplementary workbooks are byte-identical")

    if args.check_vko_data or args.check_data:
        vko_issues = audit_checksum_manifest(ROOT / "virtual_knockout/vko_input_checksums.tsv")
        if vko_issues:
            fail(errors, "virtual-knockout input checksum audit: " + ", ".join(vko_issues))
        else:
            pass_("HOA2/HOA3 virtual-knockout RDS inputs match the release SHA256 manifest")

    if args.check_data:
        data_missing = [name for name in DATA_REQUIRED if not (ROOT / name).exists()]
        gse169 = list((ROOT / "data/gse169396").glob("*")) if (ROOT / "data/gse169396").exists() else []
        gse290 = list((ROOT / "data/gse290411").glob("*")) if (ROOT / "data/gse290411").exists() else []
        if len(gse169) < 12:
            data_missing.append("data/gse169396/ (12 expected 10x files)")
        if len(gse290) < 12:
            data_missing.append("data/gse290411/ (12 expected 10x files)")
        if data_missing:
            fail(errors, "missing primary-data inputs: " + ", ".join(data_missing))
        else:
            pass_("primary-data inputs")

    if errors:
        print(f"\nRepository check failed ({len(errors)} issue groups).", file=sys.stderr)
        return 1
    print("\nRepository check passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
